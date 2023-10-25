import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Chemin du fichier Excel
excel_file_path = 'C:\\path_to_excel\\Suivi_des_backups_Veeam.xlsx'

# Chemins d'accès aux répertoires
directory1 = '\\\\192.168.x.x\\path\\Backup'
directory2 = '\\\\192.168.x.x\\path\\Backup Physique'

# Seuil pour l'espace disque restant
seuil_tb = 2.0

# Trouve le # du drive selon le fichier texte dessus
def get_value_from_files(directory_path):
    # Define the filenames to check
    filenames_to_check = ["1.txt", "2.txt", "3.txt"]

    # Regarde si un fichier existe dans le directory
    for filename in filenames_to_check:
        file_path = os.path.join(directory_path, filename)
        if os.path.exists(file_path):
            # Extrait la valeur depuis le nom de fichier (ex., "1.txt" -> "1")
            value = os.path.splitext(filename)[0]
            return value
    return None  # Return None si aucun fichier trouvé selon l'array

# Écrit la valeur dans le Excel
def mettre_a_jour_valeur_sous(feuille, ligne_actuelle, colonne):
    cellule_actuelle = feuille[f'{colonne}{ligne_actuelle}']
    cellule_au_dessus = feuille[f'{colonne}{ligne_actuelle - 1}']
    
    repo_directory = '\\\\192.168.20.76\\mtibackup'
    value = get_value_from_files(repo_directory)
    
    if value is not None:
        if value == "1":
            return 1
        elif value == "2":
            return 2
        elif value == "3":
            return 3
    else:
        return None  # Return none si aucun fichier n'est trouvé

# Fonction pour trouver la première ligne vide dans la feuille "Veeam"
def trouver_ligne_vide(feuille):
    for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, min_col=1, max_col=1):
        for cellule in ligne:
            if cellule.value is None:
                return cellule.row
    return feuille.max_row + 1

# Fonction pour obtenir l'espace utilisé d'un répertoire en téraoctets
def get_directory_space(directory_path):
    try:
        # Utilise os.walk pour parcourir récursivement le répertoire et obtenir la taille totale
        total_size_bytes = sum(os.path.getsize(os.path.join(root, filename)) for root, _, filenames in os.walk(directory_path) for filename in filenames)
        total_size_tb = total_size_bytes / (1024 ** 4)  # Convertir en téraoctets
        return round(total_size_tb, 2)  # Arrondir à 2 chiffres après la virgule
    except Exception as e:
        print(f"Erreur : {e}")
        return None

# Obtenir la date actuelle
date_actuelle = datetime.now().strftime("%d %B %Y")  # Formater la date comme désiré

# Charger le classeur Excel
classeur = openpyxl.load_workbook(excel_file_path)

# Obtenir la feuille "Veeam"
feuille = classeur['Veeam']

# Trouver la première ligne vide dans la feuille
ligne_vide = trouver_ligne_vide(feuille)

# Écrire la date actuelle dans la première ligne vide dans la colonne A
feuille[f'A{ligne_vide}'] = date_actuelle

# Écrire "Réussi" dans la deuxième colonne
feuille[f'B{ligne_vide}'] = 'Réussi'

# Obtenir l'espace utilisé des deux répertoires
space1 = get_directory_space(directory1)
space2 = get_directory_space(directory2)

if space1 is not None and space2 is not None:
    # Additionner les espaces utilisés des deux répertoires
    total_space_tb = space1 + space2

    # Calculer l'espace restant en soustrayant de 7.63 To
    remaining_space_tb = round(7.63 - total_space_tb, 2)  # Arrondir à 2 chiffres après la virgule

    # Écrire le résultat dans la quatrième colonne
    feuille[f'D{ligne_vide}'] = remaining_space_tb

    # Obtenir et mettre à jour la valeur dans la troisième colonne
    valeur_calcul = mettre_a_jour_valeur_sous(feuille, ligne_vide, 'C')
    feuille[f'C{ligne_vide}'] = valeur_calcul

    # Sauvegarder le classeur Excel mis à jour
    classeur.save(excel_file_path)

# ------------------------------- Section Email -------------------------------

# Obtenir l'espace utilisé des deux répertoires
espace1 = get_directory_space(directory1)
espace2 = get_directory_space(directory2)

if espace1 is not None and espace2 is not None:
    # Additionner les espaces utilisés des deux répertoires
    espace_total_to = espace1 + espace2

    # Calculer l'espace restant en soustrayant de 7,63 To
    espace_restant_to = round(7.63 - espace_total_to, 2)  # Arrondir à 2 chiffres après la virgule

# Vérifier si l'espace restant est inférieur au seuil
if espace_restant_to is not None and espace_restant_to < seuil_tb:
    # Configuration de l'email (remplissez les détails du serveur SMTP)
    serveur_smtp = 'abc'
    port_smtp = 587
    nom_utilisateur_smtp = 'abc'
    mot_de_passe_smtp = 'abc'
    email_expediteur = 'abc'
    email_destinataire = 'abc'

    # Créer le message email
    sujet = 'Alerte Espace Disque'
    message = f"L'espace disque restant est de {espace_restant_to:.2f} téraoctets, ce qui est inférieur au seuil de {seuil_tb} téraoctets. Ce message provient d'un script python qui s'exécute automatiquement sur le serveur xyz"

    msg = MIMEMultipart()
    msg['From'] = email_expediteur
    msg['To'] = email_destinataire
    msg['Subject'] = sujet
    msg.attach(MIMEText(message, 'plain'))

    # Se connecter au serveur SMTP et envoyer l'email
    try:
        smtp = smtplib.SMTP(serveur_smtp, port_smtp)
        smtp.starttls()
        smtp.login(nom_utilisateur_smtp, mot_de_passe_smtp)
        smtp.sendmail(email_expediteur, email_destinataire, msg.as_string())
        smtp.quit()
        print("Alerte email envoyée avec succès.")
    except Exception as e:
        print(f"Erreur lors de l'envoi de l'email : {e}")
else:
    print("L'espace disque est au-dessus du seuil. Aucune alerte envoyée.")
